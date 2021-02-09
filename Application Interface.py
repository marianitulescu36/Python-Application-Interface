import keyboard
import pyautogui
import webbrowser
import os
import openpyxl
from openpyxl import *
from openpyxl import Workbook
from time import sleep
from tkinter import *
from tkinter import messagebox
from PIL import ImageTk, Image
from tkinter.filedialog import askopenfilename
import tkinter as tk
from tkinter import ttk
from HoverInfo import HoverText # https://github.com/ImperialStranger/HoverInfo-Module

"---------------------------------- Hover Buttons -----------------------------------"
class ToolTip(object):

    def __init__(self, widget):
        self.widget = widget
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0

    def showtip(self, text):
        "Display text in tooltip window"
        self.text = text
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 20
        y = y + cy + self.widget.winfo_rooty() +10
        self.tipwindow = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = Label(tw, text=self.text, justify=LEFT,
                      background="white", relief=SOLID, borderwidth=1,
                      font=("tahoma", "9", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

def CreateToolTip(widget, text):
    toolTip = ToolTip(widget)
    def enter(event):
        toolTip.showtip(text)
    def leave(event):
        toolTip.hidetip()
    widget.bind('<Enter>', enter)
    widget.bind('<Leave>', leave)


"----------------------------------------------------------------------------------"

pathname = os.path.dirname(sys.argv[0])
cale=os.path.abspath(pathname)+'/Output Folder'

var1=0
var2=0
var3=0
var4=0


def sel1():
	global var1
	if var1==0:
		btn1.configure(image=off_btn1)
		var1=1
	else:
		btn1.configure(image=on_btn1)
		var1=0
	return var1

def sel2():
	global var2
	if var2==0:
		btn2.configure(image=off_btn2)
		var2=1
	else:
		btn2.configure(image=on_btn2)
		var2=0
	return var2

def sel3():
	global var3
	if var3==0:
		btn3.configure(image=off_btn3)
		var3=1
	else:
		btn3.configure(image=on_btn3)
		var3=0
	return var3

def sel4():
	global var4
	if var4==0:
		btn4.configure(image=off_btn4)
		var4=1
	else:
		btn4.configure(image=on_btn4)
		var4=0
	return var4


def run():
	global var1
	global var2
	global var3
	global var4
	global varchk
	check=varchk.get()

	run_btn.configure(image=off_run)

	if(var1==0 and var2==0 and var3==0 and var4==0):
		messagebox.showerror("Error","Please select an option!")
		run_btn.configure(image=on_run)
	else:
		workbook= openpyxl.Workbook()
		wb=workbook.active
		wb.title='Sheet1'

		wb.cell(row=1,column=1).value='Tab 1'
		wb.cell(row=2,column=1).value='Tab 2'
		wb.cell(row=3,column=1).value='Tab 3'
		wb.cell(row=4,column=1).value='Tab 4'
		wb.cell(row=5,column=1).value='Email'

		if var1==1:
			wb.cell(row=1,column=2).value='True'
		else:
			wb.cell(row=1,column=2).value='False'

		if var2==1:
			wb.cell(row=2,column=2).value='True'
		else:
			wb.cell(row=2,column=2).value='False'

		if var3==1:
			wb.cell(row=3,column=2).value='True'
		else:
			wb.cell(row=3,column=2).value='False'

		if var4==1:
			wb.cell(row=4,column=2).value='True'
		else:
			wb.cell(row=4,column=2).value='False'

		if check==1:
			wb.cell(row=5,column=2).value='True'
		else:
			wb.cell(row=5,column=2).value='False'

		workbook.save(str(cale)+"/OUTPUT.xlsx")

		os.startfile(pathname+"/Robot/Main.xaml")
		sleep(40)
		pyautogui.hotkey('ctrl', 'f6')

		btn1.configure(image=on_btn1)
		btn2.configure(image=on_btn2)
		btn3.configure(image=on_btn3)
		btn4.configure(image=on_btn4)
		run_btn.configure(image=on_run)
		# cb1.configure(image=off_image)
		varchk=0

app=Tk()

app.title(" ")
app.iconbitmap(default=pathname+'/Resources/white.ico')
app.geometry('502x512') 
app.configure(bg='white')



panel = Label(app, borderwidth=0, highlightthickness=0)
panel.grid(row=0, column=0, padx=20, pady=3, sticky = "NW")


title=pathname+"/Resources/Title Tab.PNG"
img_title = ImageTk.PhotoImage(Image.open(title))
panel_title = Label(app, image = img_title, borderwidth=0, highlightthickness=0)
panel_title.grid(row=1, column=0, padx=2)




path_about = os.path.dirname(sys.argv[0]) 
caleabout=os.path.abspath(path_about)+"/Output Folder/Disclaimers.pdf"
def open_disclaimer():
  webbrowser.open(caleabout)

info_icon=pathname+"/Resources/Question_Mark.PNG"
icon_about = ImageTk.PhotoImage(Image.open(info_icon))

btn_ada = Button(app, image=icon_about, fg='white', borderwidth=0, highlightthickness=0, command=open_disclaimer)
btn_ada.grid(row=0, column=0, padx=70, sticky = "E")

CreateToolTip(btn_ada, text = 'Disclaimers')


path_instr = os.path.dirname(sys.argv[0]) 
caleinstr=os.path.abspath(path_instr)+"/Output Folder/Instructions.pdf"
def open_instructions():
  webbrowser.open(caleinstr)

instr_icon=pathname+"/Resources/Info_Mark.PNG"
icon_instr = ImageTk.PhotoImage(Image.open(instr_icon))
btn_instr = Button(app, image=icon_instr, fg='white', borderwidth=0, highlightthickness=0, command=open_instructions)
btn_instr.grid(row=0, column=0, padx=25, sticky = "E")

CreateToolTip(btn_instr, text = 'Instructions')

frame=Frame(app,width=500,height=500,bg='white')
frame.grid()

btn1_on=pathname+"/Resources/Tab 1 no Shadow.PNG"
btn1_off=pathname+"/Resources/Tab 1 with Shadow.PNG"
on_btn1 = ImageTk.PhotoImage(Image.open(btn1_on))
off_btn1 = ImageTk.PhotoImage(Image.open(btn1_off))

btn2_on=pathname+"/Resources/Tab 2 no Shadow.PNG"
btn2_off=pathname+"/Resources/Tab 2 with Shadow.PNG"
on_btn2 = ImageTk.PhotoImage(Image.open(btn2_on))
off_btn2 = ImageTk.PhotoImage(Image.open(btn2_off))

btn3_on=pathname+"/Resources/Tab 3 no Shadow2.PNG"
btn3_off=pathname+"/Resources/Tab 3 with Shadow.PNG"
on_btn3 = ImageTk.PhotoImage(Image.open(btn3_on))
off_btn3 = ImageTk.PhotoImage(Image.open(btn3_off))

btn4_on=pathname+"/Resources/Tab 4 no Shadow.PNG"
btn4_off=pathname+"/Resources/Tab 4 with Shadow.PNG"
on_btn4 = ImageTk.PhotoImage(Image.open(btn4_on))
off_btn4 = ImageTk.PhotoImage(Image.open(btn4_off))

chk_on=pathname+"/Resources/Email Box with Tick.PNG"
chk_off=pathname+"/Resources/Email Box No Tick.PNG"
on_image = ImageTk.PhotoImage(Image.open(chk_on))
off_image = ImageTk.PhotoImage(Image.open(chk_off))
varchk = IntVar(value=0)

path_statistic = os.path.dirname(sys.argv[0]) 
calestat=os.path.abspath(path_statistic)+"/Output Folder/Statistics.xlsx"
def open_statistics():
  webbrowser.open(calestat)
  dash_btn.configure(image=off_dash)

dash_on=pathname+"/Resources/Dashboard No Click.PNG"
dash_off=pathname+"/Resources/Dashboard When Clicked.PNG"
on_dash = ImageTk.PhotoImage(Image.open(dash_on))
off_dash = ImageTk.PhotoImage(Image.open(dash_off))

run_on=pathname+"/Resources/Run Box No Shadow.PNG"
run_off=pathname+"/Resources/Run Box with Shadow.PNG"
on_run = ImageTk.PhotoImage(Image.open(run_on))
off_run = ImageTk.PhotoImage(Image.open(run_off))


btn1 = Button(frame, image=on_btn1, fg='white', borderwidth=0, highlightthickness=0, command=lambda: sel1())
btn1.grid(row=2, column=0, padx=2, pady=2, sticky = "SE")

btn2 = Button(frame, image=on_btn2, fg='white', borderwidth=0, highlightthickness=0, command=lambda: sel2())
btn2.grid(row=2, column=1, pady=2, sticky = "SW")

btn3 = Button(frame, image=on_btn3, fg='white', borderwidth=0, highlightthickness=0, command=lambda: sel3())
btn3.grid(row=3, column=0, padx=2, sticky = "NE")

btn4 = Button(frame, image=on_btn4, fg='white', borderwidth=0, highlightthickness=0, command=lambda: sel4())
btn4.grid(row=3, column=1, sticky = "NW")



def opt1():
    elementTypeBtn.configure(text="Company1")
    return;
def opt2():
    elementTypeBtn.configure(text="Company2")
    return;
def opt3():
    elementTypeBtn.configure(text="Company3")
    return;

elementTypeBtn = Menubutton(app, relief='raised', text='Please select a company', bg='white', width=61, pady=10, borderwidth=0, highlightthickness=0)
elementTypeBtn.grid(row=4, column=0, pady=2)
elementTypeBtn.menu = Menu(elementTypeBtn, tearoff = 0, bg='white')
elementTypeBtn.menu.add_command(label='Company1                                                                                           ', command=opt1)
elementTypeBtn.menu.add_command(label='Company2                                                                                           ', command=opt2)
elementTypeBtn.menu.add_command(label='Company3                                                                                           ', command=opt3)
elementTypeBtn['menu'] = elementTypeBtn.menu


cb1 = Checkbutton(app, image=off_image, selectimage=on_image, indicatoron=False, onvalue=1, offvalue=0, variable=varchk, borderwidth=0, highlightthickness=0)
cb1.grid(row=5, column=0, sticky = "NW", padx=1, pady=1)

dash_btn=Button(app, image=on_dash, borderwidth=0, highlightthickness=0, command=open_statistics)
dash_btn.grid(row=5, column=0,  sticky = "SW", padx=2, pady=2)

run_btn=Button(app, image=on_run, borderwidth=0, highlightthickness=0, command=run)
run_btn.grid(row=5, column=0,  sticky = "E", padx=1, pady=2)

app.mainloop()